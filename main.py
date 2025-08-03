from scraper import scrape_prices
from urllib.parse import quote
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def clean_price(price_str):
    clean_price = price_str.replace("$", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(clean_price)
    except:
        return None


def calculate_average(price_strs):
    float_prices = [clean_price(p) for p in price_strs]
    valid_prices = [p for p in float_prices if p is not None]
    if not valid_prices:
        return None
    return sum(valid_prices) / len(valid_prices)


def build_url(product_name):
    product_path = quote(product_name)
    base_url = "https://www.carrefour.com.ar/"
    url = f"{base_url}{product_path}?_q={quote(product_name)}&map=ft&order=OrderByPriceASC"
    return url


def load_products_from_txt(filename="products.txt"):
    if not os.path.exists(filename):
        print(f"File '{filename}' not found.")
        return []

    with open(filename, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]

    if not lines:
        print(f"The file '{filename}' is empty.")
        return []

    return lines


def get_products():
    print("Choose how to provide the products:")
    print("1 - Type manually")
    print("2 - Load from 'products.txt'")

    choice = input("Enter 1 or 2: ").strip()

    products = []

    if choice == "1":
        print("\nWrite in Spanish the products that you would like to research.")
        while True:
            entry = input("Type a product name or press 1 to run the research: ").strip()
            if entry == "":
                print("Empty input, please write a product name or 1 to run.")
                continue
            if entry == "1":
                if not products:
                    print("You must enter at least one product before running the research.")
                    continue
                break
            products.append(entry)
    elif choice == "2":
        products = load_products_from_txt()
        if not products:
            print("No products to process. Exiting.")
            exit()
    else:
        print("Invalid option. Exiting.")
        exit()

    return products


def main():
    products = get_products()
    filename = "results.xlsx"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Load or create workbook
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Prices"
        ws.append(["Product"])  # Header row

    # Build existing product lookup
    product_col = [cell.value for cell in ws["A"][1:]]  # Exclude header
    product_row_map = {product: idx+2 for idx, product in enumerate(product_col)}  # Row index starts at 2

    # Add new column with timestamp
    next_col = ws.max_column + 1
    ws.cell(row=1, column=next_col, value=timestamp)

    for product in products:
        url = build_url(product)
        print(f"\nSearching prices for '{product}' at: {url}")
        prices = scrape_prices(url)
        avg = calculate_average(prices)
        avg_display = f"{avg:.2f}" if avg is not None else "No prices found"
        print(f"Result for '{product}': {avg_display}")

        if product in product_row_map:
            row = product_row_map[product]
        else:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=product)
            product_row_map[product] = row

        ws.cell(row=row, column=next_col, value=avg if avg is not None else "No prices found")

    wb.save(filename)
    print(f"\nAll results saved to '{filename}'")


if __name__ == "__main__":
    main()
